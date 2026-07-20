import streamlit as st
import pandas as pd
import openpyxl
import os
import glob
import re
import copy
from datetime import datetime

# Set page config
st.set_page_config(
    page_title="Mobifone CR Generator Tool",
    page_icon="📶",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom Styling (Rich Aesthetics)
st.markdown("""
<style>
    .reportview-container {
        background: #0e1117;
    }
    .main-title {
        font-size: 38px;
        font-weight: 800;
        color: #1F4E79;
        text-align: center;
        margin-bottom: 20px;
        font-family: 'Outfit', 'Inter', sans-serif;
    }
    .subtitle {
        font-size: 16px;
        color: #8a8d93;
        text-align: center;
        margin-bottom: 40px;
    }
    .section-header {
        font-size: 22px;
        font-weight: 600;
        color: #1F4E79;
        border-bottom: 2px solid #1F4E79;
        padding-bottom: 5px;
        margin-top: 20px;
        margin-bottom: 15px;
    }
    .success-box {
        padding: 20px;
        border-radius: 10px;
        background-color: #d4edda;
        color: #155724;
        border: 1px solid #c3e6cb;
        margin-top: 15px;
    }
    .stButton>button {
        background-color: #1F4E79;
        color: white;
        font-weight: bold;
        border-radius: 8px;
        padding: 10px 25px;
        border: none;
        transition: all 0.3s ease;
    }
    .stButton>button:hover {
        background-color: #153552;
        box-shadow: 0px 4px 10px rgba(31, 78, 121, 0.4);
    }
</style>
""", unsafe_allow_html=True)

# -----------------------------------------------------------------------------
# CONSTANTS & CONFIGURATIONS
# -----------------------------------------------------------------------------
BASE_WORKSPACE_PATH = r"F:\OneDrive - Mobifone\F_WORKING\Python_Coding\Create CR"
TEMPLATE_DIR = os.path.join(BASE_WORKSPACE_PATH, "CR mẫu")
SOURCE_CELL_DIR = r"F:\OneDrive - Mobifone\F_WORKING\Python_Coding\MTCL 2026\Source_cell"
DATABASE_DIR = r"F:\OneDrive - Mobifone\F_WORKING\Python_Coding\Database trạm"

# Cấu hình 2 Netact đang hoạt động song song
NETACT_DIRS = {
    "Netact92":  {
        "4G": os.path.join(DATABASE_DIR, "4G", "Dump", "NSN", "Netact92"),
        "5G": os.path.join(DATABASE_DIR, "4G", "Dump", "NSN", "Netact92"),  # NRCELL nằm chung thư mục NSN
    },
    "Netactnh1": {
        "4G": os.path.join(DATABASE_DIR, "4G", "Dump", "NSN", "Netactnh1"),
        "5G": os.path.join(DATABASE_DIR, "4G", "Dump", "NSN", "Netactnh1"),
    },
}
# Thứ tự ưu tiên: Netactnh1 trước (dump đầy đủ hơn, mới hơn)
NETACT_PRIORITY = ["Netactnh1", "Netact92"]

# -----------------------------------------------------------------------------
# HELPER FUNCTIONS
# -----------------------------------------------------------------------------

def get_dump_mtime():
    """Trả về timestamp sửa đổi mới nhất của tất cả dump file trong cả 2 Netact.
    Dùng làm cache key: khi dump mới được copy vào, giá trị này thay đổi
    → cache tự động bị vô hiệu hóa → chương trình đọc lại dữ liệu mới.
    """
    latest = 0.0
    # Quét tất cả thư mục Netact của cả 4G và 5G
    scan_dirs = []
    for ndir in NETACT_DIRS.values():
        scan_dirs.extend(ndir.values())
    # Thêm thư mục database tổng hợp (xlsx)
    scan_dirs.append(DATABASE_DIR)

    for d in set(scan_dirs):
        if not os.path.isdir(d):
            continue
        try:
            for f in os.listdir(d):
                if f.endswith(('.csv', '.txt', '.xlsx')):
                    fp = os.path.join(d, f)
                    try:
                        t = os.path.getmtime(fp)
                        if t > latest:
                            latest = t
                    except OSError:
                        pass
        except OSError:
            pass
    return latest

@st.cache_data(show_spinner=False)
def build_cell_to_netact_map(tech, _dump_mtime=0):
    """Đọc LNCEL/NRCELL từ cả 2 Netact, trả về dict {cellName: netact_name}.
    Ưu tiên Netactnh1 (dump đầy đủ, mới nhất).
    Tham số _dump_mtime dùng làm cache key để tự động reload khi dump thay đổi.
    """
    cell_map = {}
    # Tên file dump cần tìm theo công nghệ
    dump_filenames = ["LNCEL.csv"] if tech == "4G" else ["NRCELL.csv"]
    mo_candidates  = ["$dn", "MO", "distName"]

    # Duyệt theo thứ tự ưu tiên NGƯỢC (Netact92 trước) để Netactnh1 ghi đè sau
    for netact_name in reversed(NETACT_PRIORITY):
        netact_dir = NETACT_DIRS.get(netact_name, {}).get(tech, "")
        if not os.path.isdir(netact_dir):
            continue
        for fname in dump_filenames:
            fpath = os.path.join(netact_dir, fname)
            if not os.path.exists(fpath):
                continue
            try:
                df = pd.read_csv(fpath, low_memory=False,
                                 usecols=lambda c: c in mo_candidates + ["cellName"])
                mo_col = next((c for c in mo_candidates if c in df.columns), None)
                if "cellName" not in df.columns or mo_col is None:
                    continue
                for cell_name in df["cellName"].dropna().astype(str).unique():
                    cell_name = cell_name.strip()
                    if cell_name:
                        cell_map[cell_name] = netact_name   # Netactnh1 ghi đè sau cùng → ưu tiên
            except Exception:
                pass
    return cell_map


def get_netact_dump_file(sheetname, tech, netact_name):
    """Tìm file dump (Export_*.txt hoặc sheetname.csv) trong đúng thư mục Netact.
    Trả về (filepath, is_csv) hoặc (None, False) nếu không tìm thấy.
    """
    netact_dir = NETACT_DIRS.get(netact_name, {}).get(tech, "")
    if not netact_dir or not os.path.isdir(netact_dir):
        # Fallback: thư mục NSN gốc (cho 3G hoặc legacy)
        tech_folder = "3G" if tech == "3G" else tech
        sub = "DUMP" if tech == "3G" else "Dump"
        netact_dir = os.path.join(DATABASE_DIR, tech_folder, sub, "NSN")

    # 1. Tìm Export_*.txt (ưu tiên)
    pattern_txt = os.path.join(netact_dir, f"Export_{sheetname}__*.txt")
    txt_files = glob.glob(pattern_txt)
    if not txt_files:
        # Case-insensitive fallback
        try:
            txt_files = [os.path.join(netact_dir, f)
                         for f in os.listdir(netact_dir)
                         if f.lower().startswith(f"export_{sheetname.lower()}__")]
        except OSError:
            txt_files = []
    if txt_files:
        return max(txt_files, key=os.path.getmtime), False

    # 2. Tìm sheetname.csv trong thư mục Netact
    csv_path = os.path.join(netact_dir, f"{sheetname}.csv")
    if os.path.exists(csv_path):
        return csv_path, True

    return None, False


def get_latest_cell_analysis():
    """Finds the latest Cell_Analysis_*.xlsx file > 10KB in Source_cell."""
    if not os.path.exists(SOURCE_CELL_DIR):
        return None
    files = glob.glob(os.path.join(SOURCE_CELL_DIR, "Cell_Analysis_*.xlsx"))
    valid_files = [f for f in files if os.path.getsize(f) > 10240] # > 10KB
    if not valid_files:
        return None
    return max(valid_files, key=os.path.getmtime)

@st.cache_data
def load_bad_cells_from_analysis(filepath):
    """Loads poor KPI cells from the latest Cell Analysis file."""
    bad_cells = {'3G': set(), '4G': set(), '5G': set()}
    if not filepath or not os.path.exists(filepath):
        return bad_cells
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        # 3G Top 10
        if '3G_Top10' in wb.sheetnames:
            sheet = wb['3G_Top10']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) > 1 and row[1]: # CELLBNAME NEW
                    bad_cells['3G'].add(str(row[1]).strip())
        
        # 4G Top 10
        if '4G_Top10' in wb.sheetnames:
            sheet = wb['4G_Top10']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) > 1 and row[1]: # eUTraCell New
                    bad_cells['4G'].add(str(row[1]).strip())
        
        # PRB Top 10 (Also 4G cells)
        if 'PRB_Top10' in wb.sheetnames:
            sheet = wb['PRB_Top10']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) > 1 and row[1]: # EUTRACELL
                    bad_cells['4G'].add(str(row[1]).strip())
        
        # 5G Top 10
        if '5G_Top10' in wb.sheetnames:
            sheet = wb['5G_Top10']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) > 1 and row[1]: # NRCell Name
                    bad_cells['5G'].add(str(row[1]).strip())
                    
    except Exception as e:
        st.warning(f"Không thể đọc file Cell Analysis gợi ý: {e}")
        
    return bad_cells

def extract_site_name(cell_name, tech):
    """Extracts site name from cell name based on technology."""
    s = str(cell_name).strip()
    if not s or s.lower() == 'nan':
        return ''
    if tech == '3G':
        if len(s) == 13:
            return s[:-5]  # e.g., LDGBLO07BM3GB -> LDGBLO07
        elif len(s) == 7:
            return s[:6]
        elif len(s) > 7:
            return s[:-4]
        return s[:6]
    else: # 4G & 5G
        if len(s) == 13:
            return s[:-5]  # e.g., DNIDXO09CM4CA -> DNIDXO09
        return s[:-4]

@st.cache_data
def load_cell_database(vendor, tech, _dump_mtime=0):
    """Loads or constructs cell name to distName and siteName mappings from consolidated Excel DBs.
    Includes 'netact' field for Nokia cells (Netact92 or Netactnh1).
    Tham số _dump_mtime dùng làm cache key để tự reload khi dump/database thay đổi.
    """
    mapping = {}
    # Build netact map for Nokia (4G/5G only)
    netact_map = {}
    if vendor == 'Nokia' and tech in ('4G', '5G'):
        netact_map = build_cell_to_netact_map(tech, _dump_mtime)
    try:
        # Load from consolidated database files
        filename = f"{tech}_{'NSN' if vendor == 'Nokia' else 'ERA'}.xlsx"
        db_path = os.path.join(DATABASE_DIR, filename)
        
        if os.path.exists(db_path):
            wb = openpyxl.load_workbook(db_path, read_only=True)
            if 'usual_report' in wb.sheetnames:
                df = pd.read_excel(db_path, sheet_name='usual_report')
                df.columns = [str(c).strip() for c in df.columns]
                
                # Nokia 4G mapping
                if vendor == 'Nokia' and tech == '4G':
                    # Has MRBTS, LNCEL, cellName, LNBTS
                    for _, row in df.iterrows():
                        c_name = str(row.get('cellName', '')).strip()
                        mrbts = str(row.get('MRBTS', '')).strip()
                        lnbts = str(row.get('LNBTS', '')).strip()
                        lncel = str(row.get('LNCEL', '')).strip()
                        if c_name and mrbts and lncel:
                            # LNCEL-level distName: PLMN-PLMN/MRBTS-XXXXXX/LNBTS-XXXXXX/LNCEL-XX
                            dist = f"PLMN-PLMN/{mrbts}/{lnbts}/{lncel}"
                            # LNBTS-level distName (site level): PLMN-PLMN/MRBTS-XXXXXX/LNBTS-XXXXXX
                            lnbts_dist = f"PLMN-PLMN/{mrbts}/{lnbts}"
                            site = extract_site_name(c_name, '4G')
                            mapping[c_name] = {
                                'distName': dist,
                                'lnbtsDistName': lnbts_dist,
                                'siteName': site,
                                'cellName': c_name,
                                'tech': '4G',
                                'vendor': 'Nokia',
                                'netact': netact_map.get(c_name, NETACT_PRIORITY[0]),
                            }
                # Nokia 3G mapping
                elif vendor == 'Nokia' and tech == '3G':
                    # Has distName, cellName
                    for _, row in df.iterrows():
                        c_name = str(row.get('cellName', '')).strip()
                        dist = str(row.get('distName', '')).strip()
                        if c_name and dist:
                            site = extract_site_name(c_name, '3G')
                            mapping[c_name] = {
                                'distName': dist,
                                'siteName': site,
                                'cellName': c_name,
                                'tech': '3G',
                                'vendor': 'Nokia',
                                'netact': NETACT_PRIORITY[0],  # 3G dùng Netactnh1
                            }
                # Ericsson 3G mapping
                elif vendor == 'Ericsson' and tech == '3G':
                    # CellName = short ID (S1C3), ManagedElement_id = site node (DNINTR22UL)
                    # Map both by CellName and by composite key me_id+cellName
                    for _, row in df.iterrows():
                        c_name = str(row.get('CellName', '')).strip()
                        me_id = str(row.get('ManagedElement_id', '')).strip()
                        if c_name and me_id and c_name != 'nan' and me_id != 'nan':
                            dist = f"SubNetwork=MobiFone_DongNai,MeContext={me_id},ManagedElement={me_id},UtranNetwork=1,UtranCell={c_name}"
                            entry = {
                                'distName': dist,
                                'siteName': me_id,
                                'cellName': c_name,
                                'meId': me_id,
                                'tech': '3G',
                                'vendor': 'Ericsson',
                                'netact': 'N/A',
                            }
                            mapping[c_name] = entry
                            mapping[f"{me_id}/{c_name}"] = entry
                    
                    # Also scan dump CSV for more complete data
                    dump_3g_era = os.path.join(DATABASE_DIR, '3G', 'DUMP', 'ERA', 'vsDataNodeBLocalCell.csv')
                    if os.path.exists(dump_3g_era):
                        try:
                            df_dump = pd.read_csv(dump_3g_era, low_memory=False,
                                                  usecols=['MeContext_id', 'ManagedElement_id',
                                                           'vsDataNodeBLocalCell_id'])
                            for _, drow in df_dump.iterrows():
                                me = str(drow.get('ManagedElement_id', '')).strip()
                                me_ctx = str(drow.get('MeContext_id', me)).strip()
                                cid = str(drow.get('vsDataNodeBLocalCell_id', '')).strip()
                                if me and cid and me != 'nan' and cid != 'nan':
                                    dist = f"SubNetwork=MobiFone_DongNai,MeContext={me},ManagedElement={me},UtranNetwork=1,UtranCell={cid}"
                                    entry = {
                                        'distName': dist,
                                        'siteName': me,
                                        'cellName': cid,
                                        'meId': me,
                                        'tech': '3G',
                                        'vendor': 'Ericsson',
                                        'netact': 'N/A',
                                    }
                                    comp_key = f"{me}/{cid}"
                                    mapping[comp_key] = entry
                                    # Don't overwrite xlsx-sourced entries by short key
                                    # unless not already present
                                    if cid not in mapping:
                                        mapping[cid] = entry
                        except Exception as _e:
                            pass  # dump CSV read failed, use xlsx data only
                            
                    # Also scan vsDataUtranCell.csv for full cell names (e.g. LDGXHG00BM3GA)
                    dump_3g_utran = os.path.join(DATABASE_DIR, '3G', 'DUMP', 'ERA', 'vsDataUtranCell.csv')
                    if os.path.exists(dump_3g_utran):
                        try:
                            df_utran = pd.read_csv(dump_3g_utran, low_memory=False,
                                                   usecols=['MeContext_id', 'vsDataUtranCell_id'])
                            for _, urow in df_utran.iterrows():
                                cid = str(urow.get('vsDataUtranCell_id', '')).strip()
                                me_ctx = str(urow.get('MeContext_id', '')).strip()
                                if cid and cid != 'nan':
                                    # Fallback siteName to first 8 chars if it's a long cell name, else MeContext
                                    site = cid[:8] if len(cid) >= 8 else me_ctx
                                    dist = f"SubNetwork=MobiFone_DongNai,MeContext={me_ctx},ManagedElement=1,RncFunction=1,UtranCell={cid}"
                                    entry = {
                                        'distName': dist,
                                        'siteName': site,
                                        'cellName': cid,
                                        'meId': me_ctx,
                                        'tech': '3G',
                                        'vendor': 'Ericsson',
                                        'netact': 'N/A',
                                    }
                                    if cid not in mapping:
                                        mapping[cid] = entry
                        except Exception as _e:
                            pass

                # Ericsson 4G mapping
                elif vendor == 'Ericsson' and tech == '4G':
                    # CellName = full name (DNINTR22CM4E7), ManagedElement_id = site node
                    for _, row in df.iterrows():
                        c_name = str(row.get('CellName', '')).strip()
                        me_id = str(row.get('ManagedElement_id', '')).strip()
                        if c_name and me_id:
                            dist = f"SubNetwork=MobiFone_DongNai,MeContext={me_id},ManagedElement={me_id},ENodeBFunction=1,EUtranCellFDD={c_name}"
                            entry = {
                                'distName': dist,
                                'siteName': me_id,
                                'cellName': c_name,
                                'meId': me_id,
                                'tech': '4G',
                                'vendor': 'Ericsson',
                                'netact': 'N/A',
                            }
                            mapping[c_name] = entry
                # Ericsson 5G mapping
                elif vendor == 'Ericsson' and tech == '5G':
                    # Has ManagedElement_id, CellName
                    for _, row in df.iterrows():
                        c_name = str(row.get('CellName', '')).strip()
                        me_id = str(row.get('ManagedElement_id', '')).strip()
                        if c_name and me_id:
                            dist = f"SubNetwork=MobiFone_DongNai,MeContext={me_id},ManagedElement={me_id},GNBCUCPFunction=1,NRCellCU={c_name}"
                            mapping[c_name] = {
                                'distName': dist,
                                'siteName': me_id,
                                'cellName': c_name,
                                'tech': '5G',
                                'vendor': 'Ericsson',
                                'netact': 'N/A',
                            }
            
            # Special case for Nokia 5G (where 5G_NSN.xlsx only has hardware_report)
            elif vendor == 'Nokia' and tech == '5G':
                # We can query Master_Cell_File_*.xlsx sheet '5G Cells'!
                master_files = glob.glob(os.path.join(DATABASE_DIR, "Master_Cell_File_*.xlsx"))
                if master_files:
                    latest_master = max(master_files, key=os.path.getmtime)
                    m_wb = openpyxl.load_workbook(latest_master, read_only=True)
                    if '5G Cells' in m_wb.sheetnames:
                        df_5g = pd.read_excel(latest_master, sheet_name='5G Cells')
                        df_5g.columns = [str(c).strip() for c in df_5g.columns]
                        for _, row in df_5g.iterrows():
                            c_name = str(row.get('Cell Name', '')).strip()
                            vdr = str(row.get('Vendor', '')).strip()
                            if c_name and vdr == 'NSN':
                                cell_id = int(row.get('Cell ID', 0))
                                site = extract_site_name(c_name, '5G')
                                # Construct Nokia 5G distName: PLMN-PLMN/MRBTS-XXXX/NRBTS-XXXX/NRCELL-XX
                                # Let's fetch gNodeB ID from cell info
                                gnb = str(row.get('gNodeB ID', '')).strip()
                                # e.g. MRBTS-5856010
                                if 'MRBTS-' in gnb:
                                    mrbts = gnb
                                else:
                                    mrbts = f"MRBTS-{gnb}"
                                dist = f"PLMN-PLMN/{mrbts}/NRBTS-{gnb.replace('MRBTS-', '')}/NRCELL-{cell_id}"
                                mapping[c_name] = {
                                    'distName': dist,
                                    'siteName': site,
                                    'cellName': c_name,
                                    'tech': '5G',
                                    'vendor': 'Nokia',
                                    'netact': netact_map.get(c_name, NETACT_PRIORITY[0]),
                                }
        else:
            # Fallback to scanning raw cell exports if consolidated file is missing
            st.warning(f"Không tìm thấy database tổng hợp {filename}. Sẽ quét file raw trong dump...")
            
    except Exception as e:
        st.error(f"Lỗi khi xây dựng bản đồ Cell Database: {e}")
        
    return mapping

# Cache dump file - dùng cả filepath lẫn mtime làm key để tự reload khi file thay đổi
@st.cache_data(show_spinner=False)
def load_dump_file(filepath, _mtime=0):
    """Loads a raw tab-separated text dump or CSV file and caches it.
    _mtime là modification time của file, dùng để invalidate cache khi dump được cập nhật.
    """
    try:
        if filepath.endswith('.csv'):
            return pd.read_csv(filepath, low_memory=False)
        else:
            return pd.read_csv(filepath, sep='\t', low_memory=False)
    except Exception as e:
        st.error(f"Lỗi khi đọc file dump {os.path.basename(filepath)}: {e}")
        return pd.DataFrame()


def detect_mo_class_from_dump(filepath):
    """Detects what MO class (e.g. LNBTS, LNCEL, LNHOW) the dump file represents
    by reading the first data row's MO/distName column."""
    try:
        df = load_dump_file(filepath)
        if df.empty:
            return None
        mo_col = 'MO' if 'MO' in df.columns else ('distName' if 'distName' in df.columns else ('$dn' if '$dn' in df.columns else None))
        if mo_col is None:
            return None
        sample = str(df[mo_col].dropna().iloc[0]) if not df[mo_col].dropna().empty else ''
        # Extract the last segment class name, e.g. LNBTS-470016 -> LNBTS
        parts = sample.split('/')
        if parts:
            last_part = parts[-1]  # e.g. 'LNBTS-470016' or 'LNHOW-1'
            mo_class = last_part.split('-')[0]  # e.g. 'LNBTS'
            return mo_class
    except Exception:
        pass
    return None


def get_query_dist_name(cell_info, mo_class):
    """Returns the appropriate distName for querying the dump given the MO class of the dump file.
    
    - LNBTS class  → use lnbtsDistName (site-level prefix)
    - LNCEL / LNHOW / others → use full cell distName (startswith filter)
    """
    if mo_class == 'LNBTS':
        return cell_info.get('lnbtsDistName', cell_info['distName'].rsplit('/', 1)[0])
    return cell_info['distName']

def query_nokia_parameter_value(tech, sheet_name, cell_dist_name, param_name, key_cols_values):
    """Queries the parameter value from Nokia raw export dump files."""
    tech_folder = '3G' if tech == '3G' else ('4G' if tech == '4G' else '5G')
    sub = 'DUMP' if tech == '3G' else 'Dump'
    dump_dir = os.path.join(DATABASE_DIR, tech_folder, sub, 'NSN')
    
    if not os.path.exists(dump_dir):
        return None
    
    # Locate files like Export_{sheet_name}__*.txt
    pattern = os.path.join(dump_dir, f"Export_{sheet_name}__*.txt")
    files = glob.glob(pattern)
    if not files:
        # Try case insensitive search
        all_files = os.listdir(dump_dir)
        files = [os.path.join(dump_dir, f) for f in all_files if f.lower().startswith(f"export_{sheet_name.lower()}__")]
        
    df = pd.DataFrame()
    if files:
        latest_file = max(files, key=os.path.getmtime)
        df = load_dump_file(latest_file)
    else:
        # Try CSV files under Netact92/Netactnh1 subdirectories
        netact_dirs = [os.path.join(dump_dir, 'Netact92'), os.path.join(dump_dir, 'Netactnh1')]
        dfs = []
        for d in netact_dirs:
            p = os.path.join(d, f"{sheet_name}.csv")
            if os.path.exists(p):
                df_f = load_dump_file(p)
                if not df_f.empty:
                    dfs.append(df_f)
        if dfs:
            df = pd.concat(dfs, ignore_index=True)
            
    if df.empty:
        return None
        
    # Map MO/$dn to distName if needed
    if 'MO' in df.columns:
        df = df.rename(columns={'MO': 'distName'})
    elif '$dn' in df.columns:
        df = df.rename(columns={'$dn': 'distName'})
        
    if 'distName' not in df.columns:
        return None
        
    # Filter rows by cell_dist_name
    # Associated if MO == cell_dist_name or MO starts with cell_dist_name + '/'
    df_filtered = df[df['distName'].astype(str).str.startswith(cell_dist_name)]
    if df_filtered.empty:
        return None
        
    # If other key columns exist (like utraCarrierFreq)
    for col, val in key_cols_values.items():
        if col != 'distName' and col != 'DN' and col in df_filtered.columns:
            # Try to match the key value (e.g. utraCarrierFreq = 10587)
            # Match using numeric or string depending on column type
            df_filtered = df_filtered[df_filtered[col].astype(str) == str(val)]
            
    if not df_filtered.empty and param_name in df_filtered.columns:
        return df_filtered.iloc[0][param_name]
        
    return None

def query_ericsson_parameter_value(tech, sheet_name, cell_name, cell_dist_name, param_name, key_cols_values, me_id=None):
    """Queries the parameter value from Ericsson raw CSV dumps or consolidated Excel databases.
    
    ERA CSV dump structure: columns include ManagedElement_id + vsDataXxx_id (cell id).
    Filter logic: match by ManagedElement_id AND the appropriate cell-id column.
    """
    tech_folder = '3G' if tech == '3G' else tech
    sub = 'DUMP' if tech == '3G' else 'Dump'
    dump_dir = os.path.join(DATABASE_DIR, tech_folder, sub, 'ERA')
    
    # Determine the cell ID column and value based on technology
    # 3G ERA: vsDataNodeBLocalCell_id = CellName (short, e.g. S1C3)
    # 4G ERA: vsDataEUtranCellFDD_id  = CellName (full, e.g. DNINTR22CM4E7)
    cell_id_cols_3g = ['vsDataNodeBLocalCell_id', 'vsDataNodeBSectorCarrier_id', 'CellId']
    cell_id_cols_4g = ['vsDataEUtranCellFDD_id', 'eUtranCellFDDId', 'CellName']
    cell_id_cols = cell_id_cols_3g if tech == '3G' else cell_id_cols_4g
    
    # Try raw CSV dumps first
    if os.path.exists(dump_dir):
        csv_files = []
        for name in [sheet_name, f"vsData{sheet_name}"]:
            csv_files.extend(glob.glob(os.path.join(dump_dir, f"{name}.csv")))
            csv_files.extend(glob.glob(os.path.join(dump_dir, f"{name.lower()}.csv")))
            
        if csv_files:
            latest_file = max(csv_files, key=os.path.getmtime)
            df = load_dump_file(latest_file)
            if not df.empty:
                # Filter by ManagedElement_id (site) if available
                df_filtered = df.copy()
                site_id = me_id or (cell_dist_name.split(',')[1].split('=')[1] if 'MeContext=' in cell_dist_name else None)
                if site_id and 'ManagedElement_id' in df.columns:
                    df_filtered = df_filtered[df_filtered['ManagedElement_id'].astype(str) == site_id]
                
                # Further filter by cell ID
                for cid_col in cell_id_cols:
                    if cid_col in df_filtered.columns:
                        df_cell = df_filtered[df_filtered[cid_col].astype(str) == cell_name]
                        if not df_cell.empty:
                            df_filtered = df_cell
                            break
                        
                # Filter by extra key columns
                for col, val in key_cols_values.items():
                    if col not in ['distName', 'DN', 'cellName', 'CellName', 'ManagedElement_id'] and col in df_filtered.columns:
                        df_filtered = df_filtered[df_filtered[col].astype(str) == str(val)]
                        
                if not df_filtered.empty and param_name in df_filtered.columns:
                    return df_filtered.iloc[0][param_name]
                    
    # Fallback to consolidated Excel DB (cell_report has most params, usual_report has basics)
    db_file = os.path.join(DATABASE_DIR, f"{tech}_ERA.xlsx")
    if os.path.exists(db_file):
        try:
            wb = openpyxl.load_workbook(db_file, read_only=True)
            for sheet_tab in ['cell_report', 'usual_report', 'channel_report']:
                if sheet_tab not in wb.sheetnames:
                    continue
                df = pd.read_excel(db_file, sheet_name=sheet_tab)
                df.columns = [str(c).strip() for c in df.columns]
                # Detect cell name column
                cell_col = next((c for c in ['CellName', 'cellName', 'vsDataNodeBLocalCell_id',
                                             'vsDataEUtranCellFDD_id', 'eUtranCellFDDId']
                                 if c in df.columns), None)
                if not cell_col:
                    continue
                df_filtered = df[df[cell_col].astype(str) == cell_name]
                # Also filter by ManagedElement_id to disambiguate
                if me_id and 'ManagedElement_id' in df_filtered.columns:
                    _m = df_filtered[df_filtered['ManagedElement_id'].astype(str) == me_id]
                    if not _m.empty:
                        df_filtered = _m
                if not df_filtered.empty and param_name in df_filtered.columns:
                    return df_filtered.iloc[0][param_name]
        except Exception:
            pass
            
    return None

def copy_cell_style(src_cell, dst_cell):
    """Copies all formatting styles from a source cell to a destination cell."""
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# -----------------------------------------------------------------------------
# APPLICATION MAIN INTERFACE
# -----------------------------------------------------------------------------
st.markdown("<div class='main-title'>MOBIFONE CR AUTOMATION TOOL</div>", unsafe_allow_html=True)
st.markdown("<div class='subtitle'>Tối ưu hóa quy trình cấu hình tham số mạng vô tuyến Nokia & Ericsson</div>", unsafe_allow_html=True)

# Sidebar - Trạng thái dump và nút làm mới
with st.sidebar:
    st.markdown("## ⚙️ Hệ thống")

    _current_mtime = get_dump_mtime()
    from datetime import datetime as _dt
    _mtime_str = _dt.fromtimestamp(_current_mtime).strftime("%d/%m/%Y %H:%M:%S") if _current_mtime else "Chưa có"
    st.markdown(f"**📁 Trạng thái Dump:**  \n🕒 Dump mới nhất: `{_mtime_str}`")

    for _nname in NETACT_PRIORITY:
        _ndir = NETACT_DIRS[_nname].get("4G", "")
        _latest_f, _latest_t = "", 0.0
        if os.path.isdir(_ndir):
            for _f in os.listdir(_ndir):
                if _f.endswith(('.csv', '.txt')):
                    _t = os.path.getmtime(os.path.join(_ndir, _f))
                    if _t > _latest_t:
                        _latest_t, _latest_f = _t, _f
        _icon = "🟢" if _latest_f else "🔴"
        st.markdown(f"{_icon} **{_nname}**: `{_latest_f or 'Không tìm thấy'}`")

    st.markdown("---")
    st.markdown(
        "**📌 Tự động cập nhật:**  \n"
        "Khi copy dump mới vào thư mục Netact, chương trình **tự động nhận** trong lần tải trang tiếp theo.  \n"
        "Hoặc nhấn nút bên dưới để buộc làm mới ngay."
    )
    if st.button("🔄 Làm mới dữ liệu Dump", use_container_width=True):
        st.cache_data.clear()
        st.success("✅ Đã xóa cache! Dữ liệu dump mới sẽ được tải lại.")
        st.rerun()
    st.markdown("---")
    st.caption("Mobifone CR Automation Tool v2.0")

# Layout: 2 Columns
col_left, col_right = st.columns([1, 2])


with col_left:
    st.markdown("<div class='section-header'>1. Tham số đầu vào</div>", unsafe_allow_html=True)
    
    vendor = st.selectbox("Chọn Nhà cung cấp (Vendor):", ["Nokia", "Ericsson"])
    tech = st.selectbox("Chọn Công nghệ (RAN):", ["4G", "3G", "5G"])
    
    # Input cell list or site list
    input_method = st.radio(
        "Phương thức chọn Cell/Trạm:",
        ["Gợi ý từ KPI kém SmartF", "Nhập danh sách Cell thủ công", "Nhập danh sách Trạm thủ công"]
    )
    
    selected_cell_names = []
    
    # 1. KPI kém SmartF
    if input_method == "Gợi ý từ KPI kém SmartF":
        latest_analysis = get_latest_cell_analysis()
        if latest_analysis:
            st.info(f"Đọc file KPI SmartF: **{os.path.basename(latest_analysis)}**")
            bad_cells_dict = load_bad_cells_from_analysis(latest_analysis)
            tech_bad_cells = sorted(list(bad_cells_dict.get(tech, [])))
            
            if tech_bad_cells:
                selected_cell_names = st.multiselect(
                    f"Chọn Cell lỗi {tech} ({len(tech_bad_cells)} cells gợi ý):",
                    options=tech_bad_cells,
                    default=tech_bad_cells[:5] # Default pre-select first 5
                )
            else:
                st.warning(f"Không tìm thấy Cell lỗi {tech} trong file phân tích.")
        else:
            st.warning("Không tìm thấy file Cell_Analysis_*.xlsx hợp lệ trong Source_cell.")
            
    # 2. Nhập Cell thủ công
    elif input_method == "Nhập danh sách Cell thủ công":
        cell_input = st.text_area(
            "Nhập danh sách Cell (mỗi dòng 1 cell hoặc cách nhau bởi dấu phẩy):",
            placeholder="Ví dụ: DNIDXO09CM4CA, DNIDXO09CM4CB, TGMTP1C4BA"
        )
        if cell_input:
            selected_cell_names = [c.strip() for c in re.split(r'[\n,]+', cell_input) if c.strip()]
            
    # 3. Nhập Trạm thủ công
    else:
        site_input = st.text_area(
            "Nhập danh sách Mã Trạm (mỗi dòng 1 trạm hoặc cách nhau bởi dấu phẩy):",
            placeholder="Ví dụ: DNIDXO09, TGMTP1"
        )
        if site_input:
            input_sites = [s.strip() for s in re.split(r'[\n,]+', site_input) if s.strip()]
            st.session_state['input_sites'] = input_sites
            selected_cell_names = []
        else:
            st.session_state['input_sites'] = []
            
    # CR mẫu template selection
    st.markdown("<div class='section-header'>2. Chọn CR mẫu (Template)</div>", unsafe_allow_html=True)
    if os.path.exists(TEMPLATE_DIR):
        templates = [f for f in os.listdir(TEMPLATE_DIR) if f.endswith('.xlsx')]
        if templates:
            recommended = []
            for t in templates:
                v_match = (vendor == 'Nokia' and any(k in t.lower() for k in ['nsn', 'nokia', 'nok'])) or \
                          (vendor == 'Ericsson' and any(k in t.lower() for k in ['era', 'ericsson', 'eri']))
                t_match = any(k in t.lower() for k in [tech.lower()])
                if v_match and t_match:
                    recommended.append(t)
            
            sorted_templates = recommended + [t for t in templates if t not in recommended]
            
            def format_template_option(option):
                if option in recommended:
                    return f"🌟 (Khuyên dùng) {option}"
                return option
                
            selected_template = st.selectbox(
                "Chọn file CR mẫu:",
                options=sorted_templates,
                format_func=format_template_option
            )
        else:
            st.error("Không tìm thấy file Excel mẫu nào trong thư mục CR mẫu.")
            selected_template = None
    else:
        st.error(f"Thư mục {TEMPLATE_DIR} không tồn tại.")
        selected_template = None

# Right Column - Data preview and generation
with col_right:
    st.markdown("<div class='section-header'>3. Kiểm tra thông tin & Tạo CR</div>", unsafe_allow_html=True)

    with st.spinner("Đang kết nối cơ sở dữ liệu trạm..."):
        cell_db = load_cell_database(vendor, tech, _dump_mtime=_current_mtime)
        # Load the OTHER vendor's DB too, so cell lookup works regardless of vendor
        other_vendor = 'Nokia' if vendor == 'Ericsson' else 'Ericsson'
        cell_db_other = load_cell_database(other_vendor, tech, _dump_mtime=_current_mtime)
        # Merged DB: primary vendor takes priority, fallback to other vendor
        cell_db_all = {**cell_db_other, **cell_db}

    target_cells = []

    if input_method == "Nhập danh sách Trạm thủ công":
        input_sites = st.session_state.get('input_sites', [])
        if input_sites:
            # Build a set of siteName for fast exact lookup
            site_set = set(input_sites)
            
            # Ericsson ERA: siteName = ManagedElement_id (e.g. LDGBL503, DNINTR22UL)
            # Support prefix matching: user may enter 'LDGBL503' while DB has 'LDGBL503'
            # or user enters 'DNINTR22' while DB has 'DNINTR22UL'
            def site_matches(site_name, input_set):
                if site_name in input_set:
                    return True
                # Check if any input is a prefix of siteName
                for s in input_set:
                    if site_name.startswith(s) or s.startswith(site_name):
                        return True
                return False
            
            seen_keys = set()
            for c_name, info in cell_db.items():
                sn = info.get('siteName', '')
                # Deduplicate by composite key (siteName/cellName)
                unique_key = f"{sn}/{info.get('cellName', c_name)}"
                if unique_key in seen_keys:
                    continue
                if site_matches(sn, site_set):
                    target_cells.append(info)
                    seen_keys.add(unique_key)
            if not target_cells:
                    st.warning("⚠️ Không tìm thấy trạm nào khớp trong Database. Kiểm tra lại mã trạm.")
                    if vendor == 'Ericsson':
                        st.info(
                            f"💡 Gợi ý ERA {tech}: Nhập **MeContext_id** của trạm, ví dụ **LDGBL503** hoặc **DNINTR22UL**.  \n"
                            f"Database ERA 3G hiện có {len(set(info['siteName'] for info in cell_db.values() if info.get('tech')=='3G'))} trạm. "
                            f"Một số mã: {', '.join(sorted(set(info['siteName'] for info in cell_db.values() if info.get('tech')=='3G'))[:5])}..."
                        )
                    else:
                        st.info("💡 Gợi ý Nokia: Mã trạm thường là phần đầu của tên cell, ví dụ cell **DNIDXO09CM4CA** → mã trạm **DNIDXO09**")
    else:
        # "Nhập danh sách Cell thủ công"
        # Search in ALL vendor databases (cell names are vendor-agnostic)
        not_found = []
        for c_name in selected_cell_names:
            if c_name in cell_db_all:
                target_cells.append(cell_db_all[c_name])
            else:
                # Case-insensitive fallback
                c_lower = c_name.lower()
                found = next((info for k, info in cell_db_all.items()
                              if k.lower() == c_lower), None)
                if found:
                    target_cells.append(found)
                else:
                    not_found.append(c_name)
        if not_found:
            st.warning(f"⚠️ Không tìm thấy trong Database: **{', '.join(not_found)}**. Sẽ điền trống.")
            st.info(
                f"💡 Gợi ý: Kiểm tra lại tên cell.\n"
                f"Nokia 3G/4G: ví dụ **LDGXHG28CM3EA**, **DNIDXO09CM4CA**\n"
                f"Ericsson 3G: ví dụ **DNINTR22UL/S1C3** hoặc chỉ **S1C3** (sẽ khợp trạm đầu tiên có cell đó)\n"
                f"Ericsson 4G: ví dụ **DNINTR22CM4E7**"
            )

    if target_cells:
        # ── Phân nhóm theo Netact ────────────────────────────────────────
        groups = {}
        for cell in target_cells:
            netact = cell.get('netact', NETACT_PRIORITY[0])
            groups.setdefault(netact, []).append(cell)

        # ── Hiển thị tóm tắt phân nhóm ───────────────────────────────────
        netact_names_found = sorted(groups.keys())
        has_multi_netact = len(netact_names_found) > 1

        st.write(f"Đã tìm thấy **{len(target_cells)}** cells phù hợp:")

        # Màu badge theo Netact
        NETACT_COLORS = {"Netact92": "#1565C0", "Netactnh1": "#2E7D32", "N/A": "#616161"}
        for nname in netact_names_found:
            color = NETACT_COLORS.get(nname, "#333")
            n_cells = len(groups[nname])
            st.markdown(
                f"<span style='background:{color};color:white;padding:3px 10px;"
                f"border-radius:12px;font-size:13px;margin-right:8px;'>"
                f"🔹 {nname}: {n_cells} cells</span>",
                unsafe_allow_html=True
            )

        # ── Preview theo tab ──────────────────────────────────────────────
        if has_multi_netact:
            tabs = st.tabs([f"📡 {n} ({len(groups[n])} cells)" for n in netact_names_found])
            for tab, nname in zip(tabs, netact_names_found):
                with tab:
                    preview_df = pd.DataFrame(groups[nname])
                    cols_show = [c for c in ['cellName', 'siteName', 'netact', 'distName', 'vendor', 'tech'] if c in preview_df.columns]
                    st.dataframe(preview_df[cols_show], use_container_width=True)
        else:
            preview_df = pd.DataFrame(target_cells)
            cols_show = [c for c in ['cellName', 'siteName', 'netact', 'distName', 'vendor', 'tech'] if c in preview_df.columns]
            st.dataframe(preview_df[cols_show], use_container_width=True)

        # ── Hàm tạo CR cho 1 nhóm Netact ─────────────────────────────────
        def generate_cr_for_group(group_cells, group_netact, template_path, tech, vendor):
            """Tạo file CR Excel cho một nhóm cells thuộc cùng Netact.
            Trả về (bytes_data, filename) hoặc raise Exception nếu lỗi.
            """
            wb = openpyxl.load_workbook(template_path)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_filename = f"CR_{vendor}_{tech}_{group_netact}_{timestamp}.xlsx"

            # 1. Trạm ảnh hưởng
            if 'Trạm ảnh hưởng' in wb.sheetnames:
                sheet = wb['Trạm ảnh hưởng']
                max_r = sheet.max_row
                if max_r >= 3:
                    sheet.delete_rows(3, max_r - 2)
                unique_sites = sorted(set(c['siteName'] for c in group_cells))
                for idx, site in enumerate(unique_sites):
                    row_idx = 3 + idx
                    sheet.cell(row=row_idx, column=1, value=idx + 1)
                    sheet.cell(row=row_idx, column=2, value=site)
                    sheet.cell(row=row_idx, column=3, value=tech)
                    for col in range(1, 4):
                        c = sheet.cell(row=row_idx, column=col)
                        c.font = openpyxl.styles.Font(name='Segoe UI', size=11)
                        c.alignment = openpyxl.styles.Alignment(horizontal='center' if col != 2 else 'left')
                        c.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin', color='D3D3D3'),
                            right=openpyxl.styles.Side(style='thin', color='D3D3D3'),
                            top=openpyxl.styles.Side(style='thin', color='D3D3D3'),
                            bottom=openpyxl.styles.Side(style='thin', color='D3D3D3')
                        )

            # 2. LNCELL
            if 'LNCELL' in wb.sheetnames:
                sheet = wb['LNCELL']
                max_r = sheet.max_row
                if max_r >= 2:
                    sheet.delete_rows(2, max_r - 1)
                for idx, cell in enumerate(group_cells):
                    row_idx = 2 + idx
                    sheet.cell(row=row_idx, column=1, value=cell['distName'])
                    sheet.cell(row=row_idx, column=2, value=cell['cellName'])
                    sheet.cell(row=row_idx, column=3, value=cell['siteName'])
                    for col in range(1, 4):
                        c = sheet.cell(row=row_idx, column=col)
                        c.font = openpyxl.styles.Font(name='Segoe UI', size=10)
                        c.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin', color='D3D3D3'),
                            right=openpyxl.styles.Side(style='thin', color='D3D3D3')
                        )

            # 3. Parameter sheets
            sheet_log = []
            for sheetname in wb.sheetnames:
                if sheetname in ['1. Thông tin chung', '2. Chi tiết tác động', 'Trạm ảnh hưởng', 'LNCELL']:
                    continue
                sheet = wb[sheetname]
                max_r = sheet.max_row
                if max_r < 2:
                    continue

                row_1 = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
                row_2 = [sheet.cell(row=2, column=c).value for c in range(1, sheet.max_column + 1)]

                start_old = next((i for i, v in enumerate(row_1) if v and 'old' in str(v).lower()), -1)
                start_new = next((i for i, v in enumerate(row_1) if v and 'new' in str(v).lower()), -1)
                if start_old == -1 or start_new == -1:
                    continue

                key_cols = [(row_2[i], i) for i in range(start_old)]
                params    = [row_2[i] for i in range(start_old, start_new)]
                target_new_values = {row_2[start_new + j]: sheet.cell(row=3, column=start_new + j + 1).value
                                     for j in range(len(params))}

                styles_by_col = {c: sheet.cell(row=3, column=c) for c in range(1, sheet.max_column + 1)}
                if max_r >= 3:
                    sheet.delete_rows(3, max_r - 2)

                # ── Tìm dump file từ đúng Netact ─────────────────────────
                dump_file, is_csv = get_netact_dump_file(sheetname, tech, group_netact)

                # Detect MO class từ file dump tìm được
                sheet_mo_class = None
                if dump_file:
                    sheet_mo_class = detect_mo_class_from_dump(dump_file)

                processed_lnbts = set()
                row_count = 0

                for cell in group_cells:
                    instances = []
                    # Use vendor from cell's DB entry (may differ from UI selector when mixed)
                    cell_vendor = cell.get('vendor', vendor)

                    if cell_vendor == 'Nokia' and dump_file:
                        mo_class = sheet_mo_class
                        if mo_class == 'LNBTS':
                            query_dn = get_query_dist_name(cell, mo_class)
                            if query_dn in processed_lnbts:
                                continue
                            processed_lnbts.add(query_dn)
                        else:
                            query_dn = get_query_dist_name(cell, mo_class)

                        df_dump = load_dump_file(dump_file, _mtime=os.path.getmtime(dump_file) if dump_file else 0)
                        if not df_dump.empty:
                            if '$dn' in df_dump.columns:
                                df_dump = df_dump.rename(columns={'$dn': 'distName'})
                            elif 'MO' in df_dump.columns:
                                df_dump = df_dump.rename(columns={'MO': 'distName'})
                            if 'distName' in df_dump.columns:
                                if mo_class == 'LNBTS':
                                    df_filtered = df_dump[df_dump['distName'].astype(str) == query_dn]
                                else:
                                    df_filtered = df_dump[df_dump['distName'].astype(str).str.startswith(query_dn)]
                                for _, d_row in df_filtered.iterrows():
                                    instances.append(dict(d_row))

                    elif cell_vendor == 'Ericsson':
                        # ERA dump CSV folder: 3G/DUMP/ERA  or  4G/Dump/ERA
                        tech_folder_era = '3G' if cell.get('tech', tech) == '3G' else cell.get('tech', tech)
                        sub_dir_era = 'DUMP' if tech_folder_era == '3G' else 'Dump'
                        dump_path_era = os.path.join(DATABASE_DIR, tech_folder_era, sub_dir_era, 'ERA')
                        
                        # ERA cell ID columns per technology
                        cell_id_cols_era = (
                            ['vsDataNodeBLocalCell_id', 'vsDataNodeBSectorCarrier_id', 'vsDataUtranCell_id', 'UtranCellId']
                            if tech == '3G' else
                            ['vsDataEUtranCellFDD_id', 'eUtranCellFDDId', 'CellName']
                        )
                        me_id_era = cell.get('meId', cell['siteName'])
                        
                        csv_files = []
                        if os.path.exists(dump_path_era):
                            # sheetname e.g. 'UtranCell_primaryCpichPower' -> MO class 'UtranCell'
                            # Try: exact, with vsData prefix, stripped MO class, stripped with vsData
                            mo_class_name = sheetname.split('_')[0] if '_' in sheetname else sheetname
                            candidate_names = [
                                sheetname,
                                f"vsData{sheetname}",
                                mo_class_name,
                                f"vsData{mo_class_name}",
                            ]
                            for cand in candidate_names:
                                for ext in [cand, cand.lower()]:
                                    p = os.path.join(dump_path_era, f"{ext}.csv")
                                    if os.path.exists(p) and p not in csv_files:
                                        csv_files.append(p)

                        if csv_files:
                            latest_f = max(csv_files, key=os.path.getmtime)
                            df_dump = load_dump_file(latest_f, _mtime=os.path.getmtime(latest_f))
                            if not df_dump.empty:
                                # Step 1: filter by site (MeContext_id or ManagedElement_id)
                                df_era = df_dump.copy()
                                if 'MeContext_id' in df_era.columns:
                                    # For UtranCell, me_id is usually MeContext_id (e.g. RSG091E)
                                    df_site = df_era[df_era['MeContext_id'].astype(str) == me_id_era]
                                    if not df_site.empty:
                                        df_era = df_site
                                elif 'ManagedElement_id' in df_era.columns:
                                    df_site = df_era[df_era['ManagedElement_id'].astype(str) == me_id_era]
                                    if not df_site.empty:
                                        df_era = df_site
                                # Step 2: filter by cell ID column
                                for cid_col in cell_id_cols_era:
                                    if cid_col in df_era.columns:
                                        _tmp = df_era[df_era[cid_col].astype(str) == cell['cellName']]
                                        if not _tmp.empty:
                                            df_era = _tmp
                                            break
                                for _, d_row in df_era.iterrows():
                                    instances.append(dict(d_row))
                        
                        # If no dump CSV found, fallback to ERA xlsx (cell_report/usual_report)
                        if not instances:
                            db_era = os.path.join(DATABASE_DIR, f"{tech}_ERA.xlsx")
                            if os.path.exists(db_era):
                                try:
                                    _wb_era = openpyxl.load_workbook(db_era, read_only=True)
                                    for _sh in ['cell_report', 'usual_report', 'channel_report']:
                                        if _sh not in _wb_era.sheetnames:
                                            continue
                                        _df_era = pd.read_excel(db_era, sheet_name=_sh)
                                        _df_era.columns = [str(c).strip() for c in _df_era.columns]
                                        # Find cell
                                        _cc = next((c for c in ['CellName','vsDataEUtranCellFDD_id',
                                                                  'vsDataNodeBLocalCell_id','eUtranCellFDDId']
                                                    if c in _df_era.columns), None)
                                        if not _cc:
                                            continue
                                        _df_match = _df_era[_df_era[_cc].astype(str) == cell['cellName']]
                                        if 'ManagedElement_id' in _df_match.columns:
                                            _m = _df_match[_df_match['ManagedElement_id'].astype(str) == me_id_era]
                                            if not _m.empty:
                                                _df_match = _m
                                        if not _df_match.empty:
                                            inst_row = dict(_df_match.iloc[0])
                                            inst_row['distName'] = cell['distName']
                                            inst_row['cellName'] = cell['cellName']
                                            inst_row['CellName'] = cell['cellName']
                                            instances.append(inst_row)
                                            break
                                except Exception:
                                    pass

                    if not instances:
                        eff_dist = (cell.get('lnbtsDistName', cell['distName'])
                                    if sheet_mo_class == 'LNBTS' else cell['distName'])
                        dummy = {'distName': eff_dist, 'DN': eff_dist,
                                 'cellName': cell['cellName'], 'CellName': cell['cellName'],
                                 'ManagedElement_id': cell.get('meId', cell.get('siteName', '')),
                                 'name': cell.get('siteName', '')}
                        for p_name in params:
                            if vendor == 'Nokia':
                                dummy[p_name] = query_nokia_parameter_value(tech, sheetname, eff_dist, p_name, {})
                            else:
                                me_id_fb = cell.get('meId', cell.get('siteName', ''))
                                dummy[p_name] = query_ericsson_parameter_value(
                                    tech, sheetname, cell['cellName'], cell['distName'], p_name, {}, me_id=me_id_fb
                                )
                        instances.append(dummy)

                    for inst in instances:
                        row_idx = 3 + row_count
                        row_count += 1
                        for col_name, col_offset in key_cols:
                            cell_obj = sheet.cell(row=row_idx, column=col_offset + 1)
                            copy_cell_style(styles_by_col[col_offset + 1], cell_obj)
                            if str(col_name).lower() in ['distname', 'dn']:
                                cell_obj.value = inst.get('distName', inst.get('DN', cell['distName']))
                            elif str(col_name).lower() in ['cellname', 'name']:
                                cell_obj.value = inst.get('cellName', inst.get('CellName', inst.get('name', cell['cellName'])))
                            else:
                                cell_obj.value = inst.get(col_name, '')
                        for p_idx, p_name in enumerate(params):
                            old_ci = start_old + p_idx + 1
                            new_ci = start_new + p_idx + 1
                            co = sheet.cell(row=row_idx, column=old_ci)
                            cn = sheet.cell(row=row_idx, column=new_ci)
                            copy_cell_style(styles_by_col[old_ci], co)
                            copy_cell_style(styles_by_col[new_ci], cn)
                            co.value = inst.get(p_name, None)
                            cn.value = target_new_values.get(p_name)

                sheet_log.append(f"Sheet **{sheetname}**: {row_count} dòng")

            # Lưu ra bytes
            import io as _io
            buf = _io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.read(), out_filename, sheet_log

        # ── Nút tạo CR ────────────────────────────────────────────────────
        st.markdown("---")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if st.button("🚀 Bắt đầu tạo Change Request (CR)", type="primary"):
            if not selected_template:
                st.error("Vui lòng chọn file CR mẫu trước khi tạo CR.")
            else:
                template_path = os.path.join(TEMPLATE_DIR, selected_template)
                generated_files = {}

                with st.spinner(f"Đang tạo CR cho {len(netact_names_found)} nhóm..."):
                    for nname in netact_names_found:
                        group = groups[nname]
                        try:
                            data_bytes, fname, sheet_log = generate_cr_for_group(
                                group, nname, template_path, tech, vendor
                            )
                            generated_files[nname] = (data_bytes, fname, sheet_log)
                            # Lưu ra disk
                            out_path = os.path.join(BASE_WORKSPACE_PATH, fname)
                            with open(out_path, "wb") as f:
                                f.write(data_bytes)
                        except Exception as e:
                            st.error(f"❌ Lỗi tạo CR cho {nname}: {e}")
                            st.exception(e)

                if generated_files:
                    st.success(f"✅ Tạo CR thành công cho {len(generated_files)} Netact!")
                    for nname, (data_bytes, fname, sheet_log) in generated_files.items():
                        color = NETACT_COLORS.get(nname, "#333")
                        st.markdown(
                            f"<div style='border-left:4px solid {color};padding:10px 15px;"
                            f"background:#f8f9fa;border-radius:4px;margin:10px 0;'>"
                            f"<b style='color:{color};'>📁 {nname}</b> — {len(groups[nname])} cells<br>"
                            f"<small>📄 {fname}</small><br>"
                            f"<small>{'  |  '.join(sheet_log)}</small></div>",
                            unsafe_allow_html=True
                        )
                        st.download_button(
                            label=f"📥 Tải xuống CR {nname}",
                            data=data_bytes,
                            file_name=fname,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_{nname}_{timestamp}"
                        )
    else:
        st.info("Vui lòng chọn hoặc nhập Cell/Trạm đầu vào để kiểm tra.")

